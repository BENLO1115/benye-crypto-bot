"""
Build a beautiful, feature-complete expense tracker Excel workbook.
Sheets:
  1. 儀表板 (Dashboard)  – summary cards + monthly chart
  2. 記帳明細 (Ledger)   – day-by-day entry table
  3. 月份摘要 (Monthly)  – auto pivot by month & category
  4. 預算管理 (Budget)   – budget vs actual per category
  5. 使用說明 (Guide)    – instructions
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabel, DataLabelList
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.drawing.image import Image
import openpyxl.utils.cell as cell_utils
from datetime import date, timedelta
import random

# ── Palette ────────────────────────────────────────────────────────────────────
C = {
    "navy":       "1B2A4A",
    "teal":       "0D7377",
    "teal_light": "14A085",
    "teal_pale":  "D9F2EF",
    "gold":       "F4A261",
    "gold_pale":  "FEF3E8",
    "red":        "E63946",
    "red_pale":   "FDECEA",
    "green":      "2DC653",
    "green_pale": "E6F9EC",
    "purple":     "7B2FBE",
    "purple_pale":"F3E8FF",
    "blue":       "2176FF",
    "blue_pale":  "E8F0FF",
    "gray_bg":    "F7F9FC",
    "gray_mid":   "E2E8F0",
    "gray_dark":  "64748B",
    "white":      "FFFFFF",
    "black":      "1E293B",
}

CATEGORIES = [
    "餐飲食物", "交通出行", "住宅房租", "生活日用",
    "娛樂休閒", "醫療健康", "教育學習", "服飾美容",
    "儲蓄投資", "其他支出",
]

CAT_COLORS = {
    "餐飲食物": "FF6B6B",
    "交通出行": "4ECDC4",
    "住宅房租": "45B7D1",
    "生活日用": "96CEB4",
    "娛樂休閒": "FFEAA7",
    "醫療健康": "DDA0DD",
    "教育學習": "98D8C8",
    "服飾美容": "F7DC6F",
    "儲蓄投資": "82E0AA",
    "其他支出": "AEB6BF",
}

PAYMENT_METHODS = ["現金", "信用卡", "LINE Pay", "街口支付", "悠遊卡", "銀行轉帳", "其他"]

MONTHS_ZH = ["一月","二月","三月","四月","五月","六月",
             "七月","八月","九月","十月","十一月","十二月"]

SAMPLE_DATA = [
    # (date_str, category, description, amount, payment)
    ("2026/01/02", "餐飲食物", "早餐豆漿蛋餅",    65,  "現金"),
    ("2026/01/02", "交通出行", "捷運月票",         1280, "悠遊卡"),
    ("2026/01/05", "餐飲食物", "午餐便當",         120, "LINE Pay"),
    ("2026/01/08", "娛樂休閒", "Netflix 訂閱",     270, "信用卡"),
    ("2026/01/10", "生活日用", "衛生紙洗碗精",     340, "信用卡"),
    ("2026/01/12", "醫療健康", "診所掛號費",       150, "現金"),
    ("2026/01/15", "住宅房租", "一月房租",        12000, "銀行轉帳"),
    ("2026/01/18", "教育學習", "Udemy 課程",       890, "信用卡"),
    ("2026/01/20", "服飾美容", "換季外套",        1580, "信用卡"),
    ("2026/01/22", "餐飲食物", "火鍋聚餐",         680, "LINE Pay"),
    ("2026/01/25", "儲蓄投資", "定期存款",        5000, "銀行轉帳"),
    ("2026/01/28", "其他支出", "寄快遞",            80, "現金"),
    ("2026/02/01", "餐飲食物", "年節零食",         520, "現金"),
    ("2026/02/03", "交通出行", "計程車",           180, "街口支付"),
    ("2026/02/05", "餐飲食物", "家庭年夜飯",      2200, "信用卡"),
    ("2026/02/10", "娛樂休閒", "電影票 x2",        580, "信用卡"),
    ("2026/02/12", "生活日用", "清潔用品",         420, "街口支付"),
    ("2026/02/15", "住宅房租", "二月房租",        12000, "銀行轉帳"),
    ("2026/02/18", "醫療健康", "健身房月費",       900, "信用卡"),
    ("2026/02/20", "服飾美容", "理髮",             350, "現金"),
    ("2026/02/22", "儲蓄投資", "ETF 定期定額",    3000, "銀行轉帳"),
    ("2026/02/25", "教育學習", "英文補習",        2400, "銀行轉帳"),
    ("2026/03/01", "餐飲食物", "超市採購",         860, "信用卡"),
    ("2026/03/05", "交通出行", "高鐵台北→台中",   515, "信用卡"),
    ("2026/03/08", "生活日用", "宜家家居",        1280, "信用卡"),
    ("2026/03/10", "娛樂休閒", "Switch 遊戲",     1200, "信用卡"),
    ("2026/03/12", "餐飲食物", "下午茶咖啡",       145, "LINE Pay"),
    ("2026/03/15", "住宅房租", "三月房租",        12000, "銀行轉帳"),
    ("2026/03/18", "醫療健康", "牙科洗牙",         1000, "現金"),
    ("2026/03/20", "儲蓄投資", "基金申購",        5000, "銀行轉帳"),
    ("2026/03/25", "服飾美容", "運動鞋",           2800, "信用卡"),
    ("2026/03/28", "其他支出", "禮品包裝",          120, "現金"),
]

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, size=11, color="1E293B", italic=False, name="Calibri"):
    return Font(name=name, bold=bold, size=size, color=color, italic=italic)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def border_thin(sides="all"):
    s = Side(style="thin", color="CBD5E1")
    sides_map = {
        "all":    Border(left=s, right=s, top=s, bottom=s),
        "bottom": Border(bottom=s),
        "outer":  Border(left=s, right=s, top=s, bottom=s),
    }
    return sides_map.get(sides, sides_map["all"])

def border_medium(sides="all"):
    m = Side(style="medium", color=C["teal"])
    s = Side(style="thin",   color="CBD5E1")
    if sides == "top_bottom":
        return Border(top=m, bottom=m, left=s, right=s)
    return Border(left=m, right=m, top=m, bottom=m)

def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def merge_and_write(ws, cell_range, value, f=None, p=None, a=None):
    ws.merge_cells(cell_range)
    c = ws[cell_range.split(":")[0]]
    c.value = value
    if f: c.font = f
    if p: c.fill = p
    if a: c.alignment = a
    return c

# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 1 – 儀表板
# ═══════════════════════════════════════════════════════════════════════════════
def build_dashboard(wb, sample_data):
    ws = wb.active
    ws.title = "📊 儀表板"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C["teal"]

    # Column widths
    widths = [1.5, 12, 12, 12, 12, 12, 12, 12, 12, 12, 12, 1.5]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for r in range(1, 80):
        ws.row_dimensions[r].height = 18

    # ── Header banner ──────────────────────────────────────────────────────────
    for row in range(1, 7):
        for col in range(1, 13):
            ws.cell(row=row, column=col).fill = fill(C["navy"])

    ws.row_dimensions[1].height = 8
    ws.row_dimensions[6].height = 8

    merge_and_write(ws, "B2:K3",
        "💰  個人財務記帳本",
        font(bold=True, size=22, color=C["white"], name="微軟正黑體"),
        fill(C["navy"]),
        align("center"))

    merge_and_write(ws, "B4:K4",
        "Personal Expense Tracker  |  2026 年度",
        font(size=11, color="90CAF9"),
        fill(C["navy"]),
        align("center"))

    ws.row_dimensions[5].height = 6

    # ── KPI cards (row 7-12) ───────────────────────────────────────────────────
    total_exp = sum(r[3] for r in sample_data)
    jan = sum(r[3] for r in sample_data if r[0].startswith("2026/01"))
    feb = sum(r[3] for r in sample_data if r[0].startswith("2026/02"))
    mar = sum(r[3] for r in sample_data if r[0].startswith("2026/03"))
    avg_monthly = total_exp / 3

    cards = [
        ("B", "C", "📅 本月支出",  f"NT$ {mar:,.0f}", C["teal"],      C["teal_pale"]),
        ("D", "E", "📆 月均支出",  f"NT$ {avg_monthly:,.0f}", C["blue"],  C["blue_pale"]),
        ("F", "G", "💳 最大單筆",  f"NT$ {max(r[3] for r in sample_data):,.0f}", C["purple"], C["purple_pale"]),
        ("H", "I", "🗓️ 累計支出",  f"NT$ {total_exp:,.0f}", C["gold"],   C["gold_pale"]),
        ("J", "K", "📝 記錄筆數",  f"{len(sample_data)} 筆", C["green"],  C["green_pale"]),
    ]

    ws.row_dimensions[7].height = 10
    ws.row_dimensions[8].height = 24
    ws.row_dimensions[9].height = 8
    ws.row_dimensions[10].height = 30
    ws.row_dimensions[11].height = 8
    ws.row_dimensions[12].height = 20

    for s_col, e_col, label, value, accent, bg in cards:
        rng_label = f"{s_col}8:{e_col}9"
        rng_val   = f"{s_col}10:{e_col}11"
        rng_bar   = f"{s_col}12:{e_col}12"

        ws.merge_cells(rng_label)
        ws.merge_cells(rng_val)
        ws.merge_cells(rng_bar)

        # background
        for row in range(8, 13):
            for col_letter in [s_col, e_col]:
                ws.cell(row=row, column=cell_utils.column_index_from_string(col_letter)).fill = fill(bg)

        lc = ws[f"{s_col}8"]
        lc.value = label
        lc.font  = font(bold=True, size=10, color=accent)
        lc.alignment = align("center", "bottom")
        lc.fill  = fill(bg)

        vc = ws[f"{s_col}10"]
        vc.value = value
        vc.font  = font(bold=True, size=16, color=accent)
        vc.alignment = align("center")
        vc.fill  = fill(bg)

        bc = ws[f"{s_col}12"]
        bc.fill  = fill(accent)

        # border
        for r in range(8, 13):
            for cl in [s_col, e_col]:
                ci = cell_utils.column_index_from_string(cl)
                ws.cell(row=r, column=ci).border = border_thin()

    # ── Monthly bar chart ──────────────────────────────────────────────────────
    # Write chart data in a hidden area (col M+)
    ws["M1"] = "月份"; ws["N1"] = "支出"
    data_rows = [("一月", jan), ("二月", feb), ("三月", mar)]
    for i, (m, v) in enumerate(data_rows, 2):
        ws[f"M{i}"] = m
        ws[f"N{i}"] = v

    bar = BarChart()
    bar.type = "col"
    bar.title = "每月支出趨勢 (NT$)"
    bar.style = 10
    bar.y_axis.title = "金額 (NT$)"
    bar.x_axis.title = "月份"
    bar.grouping = "clustered"
    bar.overlap = 20
    data_ref   = Reference(ws, min_col=14, min_row=1, max_row=4)
    cats_ref   = Reference(ws, min_col=13, min_row=2, max_row=4)
    bar.add_data(data_ref, titles_from_data=True)
    bar.set_categories(cats_ref)
    bar.series[0].graphicalProperties.solidFill = C["teal"]
    bar.series[0].graphicalProperties.line.solidFill = C["teal"]
    bar.width  = 14
    bar.height = 10
    ws.add_chart(bar, "B14")

    # ── Category pie chart ─────────────────────────────────────────────────────
    cat_totals = {}
    for row in sample_data:
        cat_totals[row[1]] = cat_totals.get(row[1], 0) + row[3]

    ws["P1"] = "類別"; ws["Q1"] = "金額"
    pie_rows = sorted(cat_totals.items(), key=lambda x: -x[1])
    for i, (cat, amt) in enumerate(pie_rows, 2):
        ws[f"P{i}"] = cat
        ws[f"Q{i}"] = amt

    pie = PieChart()
    pie.title = "支出類別分布"
    pie.style = 10
    pdata = Reference(ws, min_col=17, min_row=1, max_row=1+len(pie_rows))
    pcats = Reference(ws, min_col=16, min_row=2, max_row=1+len(pie_rows))
    pie.add_data(pdata, titles_from_data=True)
    pie.set_categories(pcats)
    pie.dLbls = DataLabelList()
    pie.dLbls.showPercent = True
    pie.dLbls.showCatName = False
    pie.dLbls.showVal = False
    SLICE_COLORS = ["FF6B6B","4ECDC4","45B7D1","96CEB4","FFEAA7","DDA0DD","98D8C8","F7DC6F","82E0AA","AEB6BF"]
    for idx, color in enumerate(SLICE_COLORS[:len(pie_rows)]):
        pt = DataPoint(idx=idx)
        pt.graphicalProperties.solidFill = color
        pie.series[0].dPt.append(pt)
    pie.width  = 14
    pie.height = 10
    ws.add_chart(pie, "G14")

    # ── Recent transactions table ──────────────────────────────────────────────
    start_row = 28
    merge_and_write(ws, f"B{start_row}:K{start_row}",
        "  最近消費記錄",
        font(bold=True, size=12, color=C["white"]),
        fill(C["navy"]),
        align("left", "center"))
    ws.row_dimensions[start_row].height = 26

    headers = ["日期", "類別", "說明", "付款方式", "金額 (NT$)"]
    h_cols  = ["B", "C", "D:F", "G:H", "I:K"]
    for header, col_range in zip(headers, h_cols):
        if ":" in col_range:
            rng = f"{col_range.split(':')[0]}{start_row+1}:{col_range.split(':')[1]}{start_row+1}"
            ws.merge_cells(rng)
            c = ws[f"{col_range.split(':')[0]}{start_row+1}"]
        else:
            c = ws[f"{col_range}{start_row+1}"]
        c.value = header
        c.font  = font(bold=True, size=10, color=C["white"])
        c.fill  = fill(C["teal"])
        c.alignment = align("center")
        c.border = border_thin()
    ws.row_dimensions[start_row+1].height = 22

    recent = sorted(sample_data, key=lambda x: x[0], reverse=True)[:8]
    for i, (dt, cat, desc, amt, pay) in enumerate(recent):
        row = start_row + 2 + i
        ws.row_dimensions[row].height = 20
        bg = C["white"] if i % 2 == 0 else C["gray_bg"]

        ws[f"B{row}"].value = dt
        ws[f"C{row}"].value = cat

        ws.merge_cells(f"D{row}:F{row}")
        ws[f"D{row}"].value = desc

        ws.merge_cells(f"G{row}:H{row}")
        ws[f"G{row}"].value = pay

        ws.merge_cells(f"I{row}:K{row}")
        ws[f"I{row}"].value = amt
        ws[f"I{row}"].number_format = '#,##0'
        ws[f"I{row}"].font = font(bold=True, color=C["red"] if amt >= 1000 else C["black"])

        cat_color = CAT_COLORS.get(cat, C["gray_mid"])
        for col_letter in ["B","C","D","G","I"]:
            ci = cell_utils.column_index_from_string(col_letter)
            cell = ws.cell(row=row, column=ci)
            cell.fill = fill(bg)
            cell.alignment = align("center")
            cell.border = border_thin()

    ws.column_dimensions["M"].width = 0.5
    ws.column_dimensions["N"].width = 12
    ws.column_dimensions["P"].width = 12
    ws.column_dimensions["Q"].width = 12

# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 2 – 記帳明細
# ═══════════════════════════════════════════════════════════════════════════════
def build_ledger(wb, sample_data):
    ws = wb.create_sheet("📝 記帳明細")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C["blue"]
    ws.freeze_panes = "A4"

    # Column widths
    col_cfg = [("A", 6), ("B", 14), ("C", 16), ("D", 22), ("E", 14),
               ("F", 14), ("G", 14), ("H", 20), ("I", 10)]
    for col, w in col_cfg:
        ws.column_dimensions[col].width = w

    # ── Title ──────────────────────────────────────────────────────────────────
    for col in range(1, 10):
        ws.cell(row=1, column=col).fill = fill(C["navy"])
    ws.row_dimensions[1].height = 10

    ws.merge_cells("A2:I2")
    c = ws["A2"]
    c.value = "📝  記帳明細  |  逐筆消費記錄"
    c.font  = font(bold=True, size=16, color=C["white"], name="微軟正黑體")
    c.fill  = fill(C["navy"])
    c.alignment = align("center")
    ws.row_dimensions[2].height = 32
    ws.row_dimensions[3].height = 8

    # ── Header ─────────────────────────────────────────────────────────────────
    headers = ["#", "日期", "類別", "消費說明", "金額 (NT$)", "付款方式", "帳戶", "備註", "標籤"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=i)
        c.value = h
        c.font  = font(bold=True, size=10, color=C["white"])
        c.fill  = fill(C["teal"])
        c.alignment = align("center")
        c.border = border_thin()
    ws.row_dimensions[4].height = 24

    # ── Data rows ──────────────────────────────────────────────────────────────
    sorted_data = sorted(sample_data, key=lambda x: x[0])
    for i, (dt, cat, desc, amt, pay) in enumerate(sorted_data, 1):
        row = 4 + i
        ws.row_dimensions[row].height = 20
        bg = C["white"] if i % 2 == 0 else C["gray_bg"]

        row_data = [i, dt, cat, desc, amt, pay, "玉山銀行", "", ""]
        for j, val in enumerate(row_data, 1):
            c = ws.cell(row=row, column=j, value=val)
            c.fill = fill(bg)
            c.alignment = align("center" if j in [1,2,3,5,6,7,9] else "left")
            c.border = border_thin()
            if j == 5:  # amount
                c.number_format = '#,##0'
                c.font = font(bold=True, size=10,
                              color=C["red"] if amt >= 5000 else
                              C["gold"] if amt >= 1000 else C["black"])
            else:
                c.font = font(size=10)

        # category color tag in col C
        cat_color = CAT_COLORS.get(cat, C["gray_mid"])
        ws.cell(row=row, column=3).fill = fill(cat_color + "55")  # translucent

    # ── Empty input rows ───────────────────────────────────────────────────────
    next_row = 4 + len(sorted_data) + 1
    for i in range(20):
        row = next_row + i
        ws.row_dimensions[row].height = 20
        bg = C["white"] if i % 2 == 0 else C["gray_bg"]
        # row number
        c = ws.cell(row=row, column=1, value=len(sorted_data)+i+1)
        c.fill = fill(bg); c.alignment = align("center"); c.border = border_thin()
        c.font = font(size=10, color=C["gray_dark"])
        for j in range(2, 10):
            c = ws.cell(row=row, column=j)
            c.fill = fill(bg); c.alignment = align("center"); c.border = border_thin()
            c.font = font(size=10)

    # ── Data validation – category ─────────────────────────────────────────────
    dv_cat = DataValidation(
        type="list",
        formula1='"' + ",".join(CATEGORIES) + '"',
        allow_blank=True,
        showDropDown=False,
    )
    dv_cat.error      = "請從清單選擇類別"
    dv_cat.errorTitle = "輸入錯誤"
    dv_cat.prompt     = "請選擇消費類別"
    ws.add_data_validation(dv_cat)
    dv_cat.sqref = f"C5:C{next_row+20}"

    dv_pay = DataValidation(
        type="list",
        formula1='"' + ",".join(PAYMENT_METHODS) + '"',
        allow_blank=True,
    )
    ws.add_data_validation(dv_pay)
    dv_pay.sqref = f"F5:F{next_row+20}"

    # ── Summary row ────────────────────────────────────────────────────────────
    sum_row = next_row + 21
    ws.row_dimensions[sum_row].height = 26
    ws.merge_cells(f"A{sum_row}:D{sum_row}")
    c = ws[f"A{sum_row}"]
    c.value = "📊 合計"
    c.font  = font(bold=True, size=12, color=C["white"])
    c.fill  = fill(C["navy"])
    c.alignment = align("right")
    c.border = border_thin()

    total_c = ws[f"E{sum_row}"]
    total_c.value = f"=SUM(E5:E{sum_row-1})"
    total_c.number_format = 'NT$#,##0'
    total_c.font = font(bold=True, size=12, color=C["white"])
    total_c.fill = fill(C["teal"])
    total_c.alignment = align("center")
    total_c.border = border_thin()

    for j in range(6, 10):
        c = ws.cell(row=sum_row, column=j)
        c.fill = fill(C["navy"]); c.border = border_thin()

# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 3 – 月份摘要
# ═══════════════════════════════════════════════════════════════════════════════
def build_monthly(wb, sample_data):
    ws = wb.create_sheet("📅 月份摘要")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C["purple"]

    for col in range(1, 16):
        ws.cell(row=1, column=col).fill = fill(C["navy"])
    ws.row_dimensions[1].height = 10

    ws.merge_cells("A2:O2")
    c = ws["A2"]
    c.value = "📅  月份摘要  |  各類別每月支出統計"
    c.font  = font(bold=True, size=16, color=C["white"], name="微軟正黑體")
    c.fill  = fill(C["navy"])
    c.alignment = align("center")
    ws.row_dimensions[2].height = 32
    ws.row_dimensions[3].height = 10

    # Header row
    ws.row_dimensions[4].height = 24
    ws.cell(row=4, column=1).value = "消費類別"
    ws.cell(row=4, column=1).font  = font(bold=True, size=10, color=C["white"])
    ws.cell(row=4, column=1).fill  = fill(C["navy"])
    ws.cell(row=4, column=1).alignment = align("center")
    ws.cell(row=4, column=1).border = border_thin()
    ws.column_dimensions["A"].width = 16

    months = ["一月", "二月", "三月"]
    for j, m in enumerate(months, 2):
        c = ws.cell(row=4, column=j)
        c.value = m; c.font = font(bold=True, size=10, color=C["white"])
        c.fill = fill(C["teal"]); c.alignment = align("center"); c.border = border_thin()
        ws.column_dimensions[get_column_letter(j)].width = 13

    # Total column
    ws.cell(row=4, column=5).value = "季度合計"
    ws.cell(row=4, column=5).font  = font(bold=True, size=10, color=C["white"])
    ws.cell(row=4, column=5).fill  = fill(C["gold"])
    ws.cell(row=4, column=5).alignment = align("center")
    ws.cell(row=4, column=5).border = border_thin()
    ws.column_dimensions["E"].width = 13

    # Avg column
    ws.cell(row=4, column=6).value = "月均"
    ws.cell(row=4, column=6).font  = font(bold=True, size=10, color=C["white"])
    ws.cell(row=4, column=6).fill  = fill(C["purple"])
    ws.cell(row=4, column=6).alignment = align("center")
    ws.cell(row=4, column=6).border = border_thin()
    ws.column_dimensions["F"].width = 13

    # % column
    ws.cell(row=4, column=7).value = "佔比"
    ws.cell(row=4, column=7).font  = font(bold=True, size=10, color=C["white"])
    ws.cell(row=4, column=7).fill  = fill(C["blue"])
    ws.cell(row=4, column=7).alignment = align("center")
    ws.cell(row=4, column=7).border = border_thin()
    ws.column_dimensions["G"].width = 10

    # Trend spark col
    ws.cell(row=4, column=8).value = "趨勢"
    ws.cell(row=4, column=8).font  = font(bold=True, size=10, color=C["white"])
    ws.cell(row=4, column=8).fill  = fill(C["gray_dark"])
    ws.cell(row=4, column=8).alignment = align("center")
    ws.cell(row=4, column=8).border = border_thin()
    ws.column_dimensions["H"].width = 10

    month_prefixes = ["2026/01", "2026/02", "2026/03"]
    grand_total = sum(r[3] for r in sample_data)

    for row_i, cat in enumerate(CATEGORIES, 5):
        ws.row_dimensions[row_i].height = 22
        bg = C["white"] if row_i % 2 == 0 else C["gray_bg"]
        cat_color = CAT_COLORS.get(cat, C["gray_mid"])

        c = ws.cell(row=row_i, column=1, value=cat)
        c.font = font(bold=True, size=10, color=C["navy"])
        c.fill = fill(cat_color + "44")
        c.alignment = align("center")
        c.border = border_thin()

        monthly_vals = []
        for j, prefix in enumerate(month_prefixes, 2):
            val = sum(r[3] for r in sample_data if r[0].startswith(prefix) and r[1] == cat)
            monthly_vals.append(val)
            c = ws.cell(row=row_i, column=j, value=val if val else None)
            c.number_format = '#,##0'
            c.font = font(size=10, bold=bool(val))
            c.fill = fill(bg)
            c.alignment = align("center")
            c.border = border_thin()

        total = sum(monthly_vals)
        avg   = total / 3

        tc = ws.cell(row=row_i, column=5, value=total if total else None)
        tc.number_format = 'NT$#,##0'
        tc.font = font(bold=True, size=10, color=C["gold"])
        tc.fill = fill(C["gold_pale"])
        tc.alignment = align("center"); tc.border = border_thin()

        ac = ws.cell(row=row_i, column=6, value=avg if avg else None)
        ac.number_format = 'NT$#,##0'
        ac.font = font(size=10)
        ac.fill = fill(bg); ac.alignment = align("center"); ac.border = border_thin()

        pct = total / grand_total if grand_total else 0
        pc = ws.cell(row=row_i, column=7, value=pct if pct else None)
        pc.number_format = '0.0%'
        pc.font = font(size=10)
        pc.fill = fill(bg); pc.alignment = align("center"); pc.border = border_thin()

        # Trend arrow
        if len(monthly_vals) >= 2 and monthly_vals[-1] and monthly_vals[-2]:
            arrow = "▲" if monthly_vals[-1] > monthly_vals[-2] else "▼" if monthly_vals[-1] < monthly_vals[-2] else "─"
            color = C["red"] if arrow == "▲" else C["green"] if arrow == "▼" else C["gray_dark"]
        else:
            arrow = "─"; color = C["gray_dark"]
        tc2 = ws.cell(row=row_i, column=8, value=arrow)
        tc2.font = font(bold=True, size=14, color=color)
        tc2.fill = fill(bg); tc2.alignment = align("center"); tc2.border = border_thin()

    # Total row
    total_row = 5 + len(CATEGORIES)
    ws.row_dimensions[total_row].height = 26
    for col in range(1, 9):
        ws.cell(row=total_row, column=col).fill = fill(C["navy"])
        ws.cell(row=total_row, column=col).border = border_thin()

    ws.cell(row=total_row, column=1).value = "總計"
    ws.cell(row=total_row, column=1).font  = font(bold=True, size=12, color=C["white"])
    ws.cell(row=total_row, column=1).alignment = align("center")

    for j, prefix in enumerate(month_prefixes, 2):
        mval = sum(r[3] for r in sample_data if r[0].startswith(prefix))
        c = ws.cell(row=total_row, column=j, value=mval)
        c.number_format = 'NT$#,##0'
        c.font = font(bold=True, size=11, color=C["gold"])
        c.alignment = align("center")

    c = ws.cell(row=total_row, column=5, value=grand_total)
    c.number_format = 'NT$#,##0'
    c.font = font(bold=True, size=12, color=C["gold"])
    c.alignment = align("center")

    c = ws.cell(row=total_row, column=6, value=grand_total/3)
    c.number_format = 'NT$#,##0'
    c.font = font(bold=True, size=11, color=C["gold"])
    c.alignment = align("center")

    c = ws.cell(row=total_row, column=7, value=1.0)
    c.number_format = '0.0%'
    c.font = font(bold=True, size=11, color=C["gold"])
    c.alignment = align("center")

    # Line chart
    line = LineChart()
    line.title = "各類別每月支出趨勢"
    line.style = 10
    line.y_axis.title = "金額 (NT$)"
    line.x_axis.title = "月份"
    line.width = 22; line.height = 12

    for cat_i, cat in enumerate(CATEGORIES):
        row_i = 5 + cat_i
        data_ref = Reference(ws, min_col=2, max_col=4, min_row=row_i, max_row=row_i)
        line.add_data(data_ref)

    cats_ref = Reference(ws, min_col=2, max_col=4, min_row=4, max_row=4)
    line.set_categories(cats_ref)
    ws.add_chart(line, "A20")

# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 4 – 預算管理
# ═══════════════════════════════════════════════════════════════════════════════
def build_budget(wb, sample_data):
    ws = wb.create_sheet("💡 預算管理")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C["gold"]

    BUDGETS = {
        "餐飲食物": 8000,  "交通出行": 3000, "住宅房租": 15000,
        "生活日用": 3000,  "娛樂休閒": 3000, "醫療健康": 2000,
        "教育學習": 3000,  "服飾美容": 2000, "儲蓄投資": 10000,
        "其他支出": 1000,
    }

    for col in range(1, 12):
        ws.cell(row=1, column=col).fill = fill(C["navy"])
    ws.row_dimensions[1].height = 10

    ws.merge_cells("A2:K2")
    c = ws["A2"]
    c.value = "💡  預算管理  |  預算 vs 實際支出"
    c.font  = font(bold=True, size=16, color=C["white"], name="微軟正黑體")
    c.fill  = fill(C["navy"])
    c.alignment = align("center")
    ws.row_dimensions[2].height = 32
    ws.row_dimensions[3].height = 10

    col_cfg = [("A",16),("B",14),("C",14),("D",14),("E",14),
               ("F",12),("G",16),("H",14),("I",14),("J",14),("K",14)]
    for col, w in col_cfg:
        ws.column_dimensions[col].width = w

    # Headers
    headers = ["類別","月預算(NT$)","實際支出","差額","使用率%","狀態",
               "季預算","季實際","季差額","季使用率","評級"]
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=j, value=h)
        c.font = font(bold=True, size=10, color=C["white"])
        c.fill = fill(C["teal"] if j <= 6 else C["purple"])
        c.alignment = align("center")
        c.border = border_thin()
    ws.row_dimensions[4].height = 24

    actual_by_cat = {}
    for row in sample_data:
        actual_by_cat[row[1]] = actual_by_cat.get(row[1], 0) + row[3]

    for row_i, cat in enumerate(CATEGORIES, 5):
        ws.row_dimensions[row_i].height = 22
        bg = C["white"] if row_i % 2 == 0 else C["gray_bg"]
        budget  = BUDGETS.get(cat, 1000)
        actual  = actual_by_cat.get(cat, 0) / 3  # avg monthly
        diff    = budget - actual
        pct     = actual / budget if budget else 0
        q_bgt   = budget * 3
        q_act   = actual_by_cat.get(cat, 0)
        q_diff  = q_bgt - q_act
        q_pct   = q_act / q_bgt if q_bgt else 0
        cat_color = CAT_COLORS.get(cat, C["gray_mid"])

        if pct > 1.0:   status, grade, s_color = "🔴 超支", "D", C["red"]
        elif pct > 0.85: status, grade, s_color = "🟡 接近", "C", C["gold"]
        elif pct > 0.6:  status, grade, s_color = "🟢 正常", "B", C["green"]
        else:            status, grade, s_color = "💚 良好", "A", C["teal"]

        row_vals = [cat, budget, actual, diff, pct, status,
                    q_bgt, q_act, q_diff, q_pct, grade]
        num_fmts = [None,"#,##0","#,##0","[Red]-#,##0;[Green]#,##0","0.0%",None,
                    "#,##0","#,##0","[Red]-#,##0;[Green]#,##0","0.0%",None]

        for j, (val, nfmt) in enumerate(zip(row_vals, num_fmts), 1):
            c = ws.cell(row=row_i, column=j, value=val)
            c.fill = fill(cat_color + "33") if j == 1 else fill(C["red_pale"] if pct > 1.0 and j in [3,4,5] else bg)
            c.alignment = align("center")
            c.border = border_thin()
            if nfmt: c.number_format = nfmt
            if j == 1:
                c.font = font(bold=True, size=10, color=C["navy"])
            elif j == 6:
                c.font = font(bold=True, size=10, color=s_color)
            elif j == 11:
                grade_colors = {"A": C["green"], "B": C["teal"], "C": C["gold"], "D": C["red"]}
                c.font = font(bold=True, size=12, color=grade_colors.get(grade, C["black"]))
            else:
                c.font = font(size=10)

    # Total row
    total_row = 5 + len(CATEGORIES)
    ws.row_dimensions[total_row].height = 26
    for col in range(1, 12):
        ws.cell(row=total_row, column=col).fill = fill(C["navy"])
        ws.cell(row=total_row, column=col).border = border_thin()

    total_budget = sum(BUDGETS.values())
    total_actual = sum(actual_by_cat.get(c, 0)/3 for c in CATEGORIES)
    total_diff   = total_budget - total_actual
    total_pct    = total_actual / total_budget if total_budget else 0

    vals = ["月度總計", total_budget, total_actual, total_diff, total_pct, "",
            total_budget*3, sum(actual_by_cat.values()),
            total_budget*3 - sum(actual_by_cat.values()),
            sum(actual_by_cat.values())/(total_budget*3) if total_budget else 0, ""]
    fmts = [None,"NT$#,##0","NT$#,##0","NT$#,##0","0.0%",None,
            "NT$#,##0","NT$#,##0","NT$#,##0","0.0%",None]
    for j, (val, fmt) in enumerate(zip(vals, fmts), 1):
        c = ws.cell(row=total_row, column=j, value=val)
        c.font = font(bold=True, size=11, color=C["gold"])
        c.alignment = align("center")
        if fmt: c.number_format = fmt

    # ── Budget bar chart ───────────────────────────────────────────────────────
    chart = BarChart()
    chart.type = "bar"
    chart.title = "月預算 vs 實際支出"
    chart.style = 10
    chart.y_axis.title = "金額 (NT$)"
    chart.x_axis.title = "類別"
    chart.grouping = "clustered"
    chart.width = 24; chart.height = 14

    budget_ref = Reference(ws, min_col=2, max_col=3,
                           min_row=4, max_row=4+len(CATEGORIES))
    cats_ref   = Reference(ws, min_col=1, min_row=5, max_row=4+len(CATEGORIES))
    chart.add_data(budget_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.solidFill = C["teal"]
    chart.series[1].graphicalProperties.solidFill = C["red"]
    ws.add_chart(chart, "A20")

# ═══════════════════════════════════════════════════════════════════════════════
# Sheet 5 – 使用說明
# ═══════════════════════════════════════════════════════════════════════════════
def build_guide(wb):
    ws = wb.create_sheet("📖 使用說明")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = C["gray_dark"]

    col_cfg = [("A",4),("B",26),("C",50),("D",4)]
    for col, w in col_cfg:
        ws.column_dimensions[col].width = w

    for col in range(1, 5):
        ws.cell(row=1, column=col).fill = fill(C["navy"])
    ws.row_dimensions[1].height = 10

    ws.merge_cells("B2:C2")
    c = ws["B2"]
    c.value = "📖  使用說明  |  User Guide"
    c.font  = font(bold=True, size=18, color=C["white"], name="微軟正黑體")
    c.fill  = fill(C["navy"])
    c.alignment = align("center")
    ws.row_dimensions[2].height = 36
    ws.row_dimensions[3].height = 12

    sections = [
        ("工作表說明", [
            ("📊 儀表板",  "財務總覽：KPI 摘要卡、月支出長條圖、類別圓餅圖、最近消費"),
            ("📝 記帳明細","逐筆手動輸入支出；類別與付款方式有下拉選單輔助輸入"),
            ("📅 月份摘要","依月份×類別交叉統計，附趨勢箭頭與佔比"),
            ("💡 預算管理","設定月預算，自動計算使用率與超支狀態評級"),
        ]),
        ("如何新增記帳", [
            ("Step 1", "切換到『📝 記帳明細』工作表"),
            ("Step 2", "在最後一筆空白行輸入：日期（格式 YYYY/MM/DD）"),
            ("Step 3", "點選「類別」欄位，從下拉選單選擇消費類別"),
            ("Step 4", "填入消費說明、金額（純數字）、付款方式"),
            ("Step 5", "備註欄可記錄發票號碼或其他資訊"),
        ]),
        ("顏色規則說明", [
            ("金額標色",  "NT$5,000 以上：紅色字體 | NT$1,000 以上：橙色字體"),
            ("狀態標示",  "💚 良好(≤60%) | 🟢 正常(≤85%) | 🟡 接近(≤100%) | 🔴 超支(>100%)"),
            ("趨勢箭頭",  "▲ 紅色 = 支出上升 | ▼ 綠色 = 支出下降 | ─ 持平"),
        ]),
        ("快捷技巧", [
            ("Ctrl+;",     "快速輸入今天日期"),
            ("Alt+↓",      "在下拉選單欄位開啟清單"),
            ("Ctrl+Shift+$", "快速套用貨幣格式"),
            ("F2",         "進入儲存格編輯模式"),
        ]),
    ]

    current_row = 4
    for section_title, items in sections:
        ws.row_dimensions[current_row].height = 28
        ws.merge_cells(f"B{current_row}:C{current_row}")
        c = ws[f"B{current_row}"]
        c.value = f"  {section_title}"
        c.font  = font(bold=True, size=13, color=C["white"])
        c.fill  = fill(C["teal"])
        c.alignment = align("left", "center")
        c.border = border_thin()
        current_row += 1

        for key, val in items:
            ws.row_dimensions[current_row].height = 22
            ws.cell(row=current_row, column=2).value = f"  {key}"
            ws.cell(row=current_row, column=2).font  = font(bold=True, size=10, color=C["navy"])
            ws.cell(row=current_row, column=2).fill  = fill(C["teal_pale"])
            ws.cell(row=current_row, column=2).alignment = align("left", "center")
            ws.cell(row=current_row, column=2).border = border_thin()

            ws.cell(row=current_row, column=3).value = f"  {val}"
            ws.cell(row=current_row, column=3).font  = font(size=10, color=C["black"])
            ws.cell(row=current_row, column=3).fill  = fill(C["gray_bg"])
            ws.cell(row=current_row, column=3).alignment = align("left", "center")
            ws.cell(row=current_row, column=3).border = border_thin()
            current_row += 1

        current_row += 1

    # Footer
    ws.row_dimensions[current_row+1].height = 30
    ws.merge_cells(f"B{current_row+1}:C{current_row+1}")
    c = ws[f"B{current_row+1}"]
    c.value = "💰  個人財務記帳本  |  設計精良，記帳輕鬆  |  祝您財務健康 🎉"
    c.font  = font(bold=True, size=12, color=C["white"])
    c.fill  = fill(C["navy"])
    c.alignment = align("center")

# ═══════════════════════════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    wb = Workbook()
    build_dashboard(wb, SAMPLE_DATA)
    build_ledger(wb, SAMPLE_DATA)
    build_monthly(wb, SAMPLE_DATA)
    build_budget(wb, SAMPLE_DATA)
    build_guide(wb)

    # Set opening sheet to dashboard
    wb.active = wb["📊 儀表板"]

    out = "/home/user/benye-crypto-bot/個人財務記帳本.xlsx"
    wb.save(out)
    print(f"✅ Saved: {out}")

if __name__ == "__main__":
    main()
